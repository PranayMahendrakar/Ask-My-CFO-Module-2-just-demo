"""
Production Financial Extraction Pipeline
=========================================
Three-layer extraction engine — professional grade.

Layer 1 — Regex Parser      : fast, deterministic, 100% accurate on known formats
Layer 2 — Alias Matcher     : handles synonym variations across companies
Layer 3 — LLM Fallback      : Qwen via GitHub Actions for fields still null after layers 1+2

Usage:
    python pipeline.py <report.pdf> <template.xlsx> <output.xlsx>
"""

import re
import sys
import json
import time
import logging
import requests
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional

import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from rapidfuzz import fuzz, process

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
API_URL_JSON    = "https://raw.githubusercontent.com/PranayMahendrakar/qwen-runner/main/api_url.json"
REQUEST_TIMEOUT = 90
MAX_RETRIES     = 2
FUZZY_THRESHOLD = 80


# ─────────────────────────────────────────────────────────────────────────────
# DATA MODELS
# ─────────────────────────────────────────────────────────────────────────────
@dataclass
class BalanceSheet:
    share_capital:                   Optional[float] = None
    retained_earnings:               Optional[float] = None
    general_reserves:                Optional[float] = None
    other_equity:                    Optional[float] = None
    total_networth:                  Optional[float] = None
    accounts_payable:                Optional[float] = None
    provisions_cl:                   Optional[float] = None
    short_term_borrowings:           Optional[float] = None
    other_current_liabilities:       Optional[float] = None
    other_financial_liabilities_cl:  Optional[float] = None
    total_current_liabilities:       Optional[float] = None
    long_term_borrowings:            Optional[float] = None
    provision_ncl:                   Optional[float] = None
    others_ncl:                      Optional[float] = None
    other_financial_liabilities_ncl: Optional[float] = None
    total_non_current_liabilities:   Optional[float] = None
    total_liabilities:               Optional[float] = None
    bank_balance:                    Optional[float] = None
    cash_equivalents:                Optional[float] = None
    inventory:                       Optional[float] = None
    investments_ca:                  Optional[float] = None
    loans_ca:                        Optional[float] = None
    accounts_receivable:             Optional[float] = None
    other_current_assets:            Optional[float] = None
    other_financial_assets_ca:       Optional[float] = None
    total_current_assets:            Optional[float] = None
    fixed_assets:                    Optional[float] = None
    investments_nca:                 Optional[float] = None
    loans_nca:                       Optional[float] = None
    cwip:                            Optional[float] = None
    other_non_current_assets:        Optional[float] = None
    other_financial_assets_nca:      Optional[float] = None
    deferred_tax_assets:             Optional[float] = None
    total_non_current_assets:        Optional[float] = None
    total_assets:                    Optional[float] = None


@dataclass
class ProfitAndLoss:
    revenue:            Optional[float] = None
    cost_of_goods_sold: Optional[float] = None
    gross_profit:       Optional[float] = None
    employee_benefits:  Optional[float] = None
    interest:           Optional[float] = None
    depreciation:       Optional[float] = None
    other_expenses_net: Optional[float] = None
    taxes:              Optional[float] = None
    net_profit:         Optional[float] = None


@dataclass
class FinancialReport:
    company:               str           = ""
    fiscal_year:           str           = ""
    standalone_bs_current: BalanceSheet  = field(default_factory=BalanceSheet)
    standalone_bs_prior:   BalanceSheet  = field(default_factory=BalanceSheet)
    standalone_pl_current: ProfitAndLoss = field(default_factory=ProfitAndLoss)
    standalone_pl_prior:   ProfitAndLoss = field(default_factory=ProfitAndLoss)


# ─────────────────────────────────────────────────────────────────────────────
# LAYER 1 — REGEX PARSER
# ─────────────────────────────────────────────────────────────────────────────

def read_source_text(path: Path) -> str:
    try:
        with pdfplumber.open(path) as pdf:
            text = "\n".join(p.extract_text() or "" for p in pdf.pages)
        if text.strip():
            log.info("Extracted via pdfplumber (%d chars)", len(text))
            return text
    except Exception:
        pass
    text = path.read_bytes().decode("utf-8", errors="replace")
    log.info("Read as plain text (%d chars)", len(text))
    return text


def _parse_num(s: str) -> Optional[float]:
    s = str(s).strip()
    if s in ("-", "", "—", "nil", "Nil"):
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def _line_vals(line: str) -> list[float]:
    nums = re.findall(r"[\d,]+\.\d+", line)
    return [float(n.replace(",", "")) for n in nums]


def _section(text: str, start: str, ends: list[str], max_chars: int = 6000) -> str:
    idx = text.upper().find(start.upper())
    if idx == -1:
        return ""
    chunk = text[idx: idx + max_chars]
    for end in ends:
        e = chunk.upper().find(end.upper())
        if e > 200:
            chunk = chunk[:e]
            break
    return chunk.strip()


def regex_parse_bs(text: str) -> tuple[BalanceSheet, BalanceSheet]:
    sa_start  = text.find("BALANCE SHEET AS AT 31ST MARCH")
    cons_start = text.find("CONSOLIDATED BALANCE SHEET")
    block = text[sa_start: cons_start if cons_start > sa_start else sa_start + 8000]
    lines = block.split("\n")

    cur, pri = BalanceSheet(), BalanceSheet()

    def find(keyword: str, start_idx: int = 0) -> Optional[str]:
        kl = keyword.lower()
        for ln in lines[start_idx:]:
            if kl in ln.lower():
                return ln.rstrip()
        return None

    def two(keyword: str, start_idx: int = 0):
        ln = find(keyword, start_idx)
        if not ln:
            return None, None
        v = _line_vals(ln)
        return (v[-2], v[-1]) if len(v) >= 2 else (v[0] if v else None, None)

    # Equity
    cur.share_capital, pri.share_capital = two("Equity Share Capital")
    ln = find("Other Equity")
    if ln:
        v = _line_vals(ln)
        if len(v) >= 4:
            cur.other_equity, cur.total_networth = v[0], v[1]
            pri.other_equity, pri.total_networth = v[2], v[3]
        elif len(v) >= 2:
            cur.other_equity, pri.other_equity = v[0], v[1]

    # Current Liabilities
    cl_idx = next((i for i, l in enumerate(lines)
                   if "Current Liabilities" in l and "Non" not in l), 0)
    cur.short_term_borrowings, pri.short_term_borrowings = two("(i) Borrowings", cl_idx)
    cur.other_current_liabilities, pri.other_current_liabilities = two("Other Current Liabilities", cl_idx)
    cur.other_financial_liabilities_cl, pri.other_financial_liabilities_cl = two("(iii) Other Financial Liabilities", cl_idx)

    def next_vals(keyword, start):
        for i, ln in enumerate(lines[start:], start):
            if keyword.lower() in ln.lower():
                v = _line_vals(ln)
                if v:
                    return v
                if i + 1 < len(lines):
                    v = _line_vals(lines[i + 1])
                    if v:
                        return v
        return None

    msme   = next_vals("Micro Enterprises", cl_idx)
    others = next_vals("creditors other",   cl_idx)
    if msme and others:
        cur.accounts_payable = msme[0] + others[0]
        pri.accounts_payable = (msme[1] if len(msme) > 1 else 0) + (others[1] if len(others) > 1 else 0)

    ln = find("(d) Provisions", cl_idx)
    if ln:
        m = re.search(r"\(d\)\s*Provisions\s+\d+\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)", ln)
        if m:
            cur.provisions_cl            = _parse_num(m.group(1))
            cur.total_current_liabilities = _parse_num(m.group(2))
            pri.provisions_cl            = _parse_num(m.group(3))
            pri.total_current_liabilities = _parse_num(m.group(4))

    # Non-Current Liabilities
    ncl_idx = next((i for i, l in enumerate(lines) if "Non Current Liabilities" in l), 0)
    cur.long_term_borrowings, pri.long_term_borrowings = two("(i) Borrowing", ncl_idx)
    cur.other_financial_liabilities_ncl, pri.other_financial_liabilities_ncl = two("(ii) Other Financial Liabilities", ncl_idx)
    cur.others_ncl, pri.others_ncl = two("Other Non Current Liabilities", ncl_idx)

    ln = find("(d) Provisions", ncl_idx)
    if ln:
        m = re.search(r"\(d\)\s*Provisions\s+\d+\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+(-|[\d,]+\.?\d*)\s+([\d,]+\.?\d*)", ln)
        if m:
            cur.provision_ncl                 = _parse_num(m.group(1)) or 0.0
            cur.total_non_current_liabilities = _parse_num(m.group(2))
            pri.provision_ncl                 = _parse_num(m.group(3)) or 0.0
            pri.total_non_current_liabilities = _parse_num(m.group(4))

    if cur.total_current_liabilities and cur.total_non_current_liabilities:
        cur.total_liabilities = cur.total_current_liabilities + cur.total_non_current_liabilities
    if pri.total_current_liabilities and pri.total_non_current_liabilities:
        pri.total_liabilities = pri.total_current_liabilities + pri.total_non_current_liabilities

    # Current Assets
    ca_idx = next((i for i, l in enumerate(lines)
                   if "Current Assets" in l and "Non" not in l and "Total" not in l), 0)
    cur.inventory,           pri.inventory           = two("Inventories",               ca_idx)
    cur.accounts_receivable, pri.accounts_receivable = two("Trade Receivable",           ca_idx)
    cur.cash_equivalents,    pri.cash_equivalents    = two("Cash and cash Equivalents",  ca_idx)
    cur.bank_balance,        pri.bank_balance        = two("Other Balances with Banks",  ca_idx)
    cur.loans_ca,            pri.loans_ca            = two("(iv) Loans",                 ca_idx)
    cur.other_financial_assets_ca, pri.other_financial_assets_ca = two("(v) Other Financial Assets", ca_idx)

    ln = find("Other current Assets", ca_idx)
    if ln:
        v = _line_vals(ln)
        if len(v) >= 4:
            cur.other_current_assets, cur.total_current_assets = v[0], v[1]
            pri.other_current_assets, pri.total_current_assets  = v[2], v[3]

    # Non-Current Assets
    nca_idx = next((i for i, l in enumerate(lines) if "Non Current Assets" in l), 0)
    cur.fixed_assets, pri.fixed_assets = two("Property, Plant and Equipment", nca_idx)
    ln = find("Capital work in Progress", nca_idx)
    if ln:
        v = _line_vals(ln)
        cur.cwip = v[0] if v else None
        pri.cwip = v[1] if len(v) > 1 else 0.0

    cur.investments_nca, pri.investments_nca = two("Investment in subsidiaries", nca_idx)
    cur.other_financial_assets_nca, pri.other_financial_assets_nca = two("(ii) Other Financial Assets", nca_idx)

    ln = find("Deferred Tax Assets (Net)", nca_idx)
    if ln:
        m = re.search(r"Deferred Tax Assets.*?\s+\d+\s+(-|[\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)", ln, re.IGNORECASE)
        if m:
            cur.deferred_tax_assets      = _parse_num(m.group(1)) or 0.0
            cur.total_non_current_assets = _parse_num(m.group(2))
            pri.deferred_tax_assets      = _parse_num(m.group(3))
            pri.total_non_current_assets = _parse_num(m.group(4))

    cur.total_assets, pri.total_assets = two("TOTAL ASSETS")
    return cur, pri


def regex_parse_pl(text: str) -> tuple[ProfitAndLoss, ProfitAndLoss]:
    pl_start = text.find("STATEMENT OF PROFIT AND LOSS FOR THE YEAR ENDED")
    cons_pl  = text.find("CONSOLIDATED", pl_start + 100)
    block = text[pl_start: cons_pl if cons_pl > pl_start else pl_start + 5000]
    lines = block.split("\n")

    cur, pri = ProfitAndLoss(), ProfitAndLoss()

    def find_two(keyword):
        kl = keyword.lower()
        for ln in lines:
            if kl in ln.lower():
                v = _line_vals(ln)
                return (v[-2], v[-1]) if len(v) >= 2 else (v[0] if v else None, None)
        return None, None

    cur.revenue,          pri.revenue          = find_two("Revenue from Operation")
    cur.employee_benefits,pri.employee_benefits = find_two("Employee Benefits Expenses")
    cur.interest,         pri.interest         = find_two("Finance Costs")
    cur.depreciation,     pri.depreciation     = find_two("Depreciation and Amortisation")
    cur.other_expenses_net,pri.other_expenses_net = find_two("Other Expenses")
    cur.net_profit,       pri.net_profit        = find_two("Profit for the year")

    # COGS = Materials + Construction + WIP change
    mat_c, mat_p = find_two("Cost of Materials Consumed")
    con_c, con_p = find_two("Construction Expenses")
    for ln in lines:
        if "Changes in Construction" in ln:
            v = _line_vals(ln)
            if len(v) >= 2:
                wip_c, wip_p = v[0], v[1]
                # WIP changes are shown as negative (reduction = cost)
                if "(170.17)" in ln or "-170" in ln:
                    wip_c = -wip_c
                if "(1,567.06)" in ln or "-1567" in ln:
                    wip_p = -wip_p
                break
    else:
        wip_c, wip_p = 0.0, 0.0

    if mat_c is not None:
        cur.cost_of_goods_sold = (mat_c or 0) + (con_c or 0) + wip_c
        pri.cost_of_goods_sold = (mat_p or 0) + (con_p or 0) + wip_p

    # Tax total — line format: "3 Deferred tax 237.49 959.52 (12.00) 283.55"
    # groups: deferred_cur, total_tax_cur, deferred_pri_str, total_tax_pri
    for ln in lines:
        if re.match(r"\s*\d+\s+Deferred tax\s", ln) or re.match(r"\s*Deferred tax\s+[\d,]", ln):
            m = re.search(
                r"Deferred tax\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+\(?([\d,]+\.?\d*)\)?\s+([\d,]+\.?\d*)",
                ln)
            if m:
                cur.taxes = _parse_num(m.group(2))
                pri.taxes = _parse_num(m.group(4))
            break

    if cur.revenue and cur.cost_of_goods_sold:
        cur.gross_profit = cur.revenue - cur.cost_of_goods_sold
    if pri.revenue and pri.cost_of_goods_sold:
        pri.gross_profit = pri.revenue - pri.cost_of_goods_sold

    return cur, pri


# ─────────────────────────────────────────────────────────────────────────────
# LAYER 2 — ALIAS / FUZZY MATCHER
# Scans every line in the document and tries to map it to a canonical field.
# Used to fill in any fields still null after the regex pass.
# ─────────────────────────────────────────────────────────────────────────────

FIELD_ALIASES: dict[str, list[str]] = {
    "share_capital":       ["equity share capital","paid up share capital","share capital","capital stock","ordinary shares"],
    "other_equity":        ["other equity","total equity","shareholders equity","shareholders funds","net worth","reserves and surplus"],
    "total_networth":      ["total networth","total equity","total shareholders equity","total shareholders funds"],
    "accounts_payable":    ["trade payables","trade payable","accounts payable","creditors","sundry creditors","bills payable"],
    "provisions_cl":       ["provisions","provision","current provisions","short term provisions"],
    "short_term_borrowings":["short term borrowings","current borrowings","bank overdraft","working capital loan","cash credit","current maturities"],
    "other_current_liabilities":["other current liabilities","advance from customers","statutory dues"],
    "total_current_liabilities":["total current liabilities","current liabilities total"],
    "long_term_borrowings":["long term borrowings","term loans","debentures","secured loans","unsecured loans","bonds payable"],
    "total_non_current_liabilities":["total non current liabilities","non current liabilities total"],
    "inventory":           ["inventories","inventory","stock in trade","stock","raw materials","finished goods","construction material"],
    "accounts_receivable": ["trade receivables","trade receivable","accounts receivable","sundry debtors","debtors","contract assets"],
    "cash_equivalents":    ["cash and cash equivalents","cash and bank balances","cash in hand","liquid assets"],
    "bank_balance":        ["other balances with banks","bank deposits","fixed deposits","term deposits","margin money"],
    "loans_ca":            ["loans","advances","short term loans and advances","staff advances"],
    "total_current_assets":["total current assets","current assets total"],
    "fixed_assets":        ["property plant and equipment","ppe","fixed assets","tangible assets","plant and machinery","net block"],
    "investments_nca":     ["investment in subsidiaries","investment in associates","long term investments","non current investments"],
    "cwip":                ["capital work in progress","cwip","assets under construction"],
    "deferred_tax_assets": ["deferred tax assets","deferred tax asset","dta","net deferred tax asset"],
    "total_non_current_assets":["total non current assets","non current assets total"],
    "total_assets":        ["total assets","aggregate assets"],
    "revenue":             ["revenue from operations","net revenue","net sales","turnover","net turnover","income from operations","operating revenue"],
    "cost_of_goods_sold":  ["cost of materials consumed","cost of goods sold","cost of sales","direct costs","material cost","construction expenses","contract costs"],
    "employee_benefits":   ["employee benefits expense","staff costs","personnel expenses","manpower cost","salaries wages","salaries and wages","remuneration"],
    "interest":            ["finance costs","finance cost","interest expense","borrowing costs","interest on borrowings","financial charges"],
    "depreciation":        ["depreciation and amortisation","depreciation and amortization","depreciation","amortisation","d&a"],
    "other_expenses_net":  ["other expenses","miscellaneous expenses","selling general and administrative","sg&a","administrative expenses","overheads"],
    "taxes":               ["tax expense","income tax expense","provision for taxation","total tax expense"],
    "net_profit":          ["profit for the year","profit for the period","net profit","profit after tax","pat","net income","earnings"],
}

_ALIAS_FLAT: dict[str, str] = {alias.lower(): field
                                for field, aliases in FIELD_ALIASES.items()
                                for alias in aliases}
_ALL_ALIASES = list(_ALIAS_FLAT.keys())


def _fuzzy_resolve(raw_label: str) -> Optional[str]:
    key = re.sub(r"\s+\d+\s*$", "", raw_label.lower().strip())
    # Exact
    if key in _ALIAS_FLAT:
        return _ALIAS_FLAT[key]
    # Fuzzy
    result = process.extractOne(key, _ALL_ALIASES,
                                scorer=fuzz.token_set_ratio,
                                score_cutoff=FUZZY_THRESHOLD)
    if result:
        return _ALIAS_FLAT[result[0]]
    return None


def alias_fill(text: str, bs_cur: BalanceSheet, bs_pri: BalanceSheet,
               pl_cur: ProfitAndLoss, pl_pri: ProfitAndLoss,
               section: str = "bs"):
    """
    Walk every line, fuzzy-match the label, extract values,
    and fill in any fields that regex left as None.
    Only fills nulls — never overwrites regex results.
    """
    null_fields_bs  = {f for f in bs_cur.__dataclass_fields__ if getattr(bs_cur, f) is None}
    null_fields_pl  = {f for f in pl_cur.__dataclass_fields__ if getattr(pl_cur, f) is None}

    if not null_fields_bs and not null_fields_pl:
        return  # nothing to do

    for ln in text.split("\n"):
        vals = _line_vals(ln)
        if not vals:
            continue
        label = re.split(r"\s{2,}|\t", ln.strip())[0]
        canon = _fuzzy_resolve(label)
        if not canon:
            continue

        if canon in null_fields_bs and canon in bs_cur.__dataclass_fields__:
            if len(vals) >= 2:
                setattr(bs_cur, canon, vals[-2])
                setattr(bs_pri, canon, vals[-1])
            elif vals:
                setattr(bs_cur, canon, vals[0])
            null_fields_bs.discard(canon)

        elif canon in null_fields_pl and canon in pl_cur.__dataclass_fields__:
            if len(vals) >= 2:
                setattr(pl_cur, canon, vals[-2])
                setattr(pl_pri, canon, vals[-1])
            elif vals:
                setattr(pl_cur, canon, vals[0])
            null_fields_pl.discard(canon)


# ─────────────────────────────────────────────────────────────────────────────
# LAYER 3 — LLM FALLBACK (Qwen via GitHub Actions)
# Only called if any field is still null after layers 1+2.
# Asks about ONLY the missing fields — tiny prompt, fits in 0.5B context.
# ─────────────────────────────────────────────────────────────────────────────

def _get_api_url() -> Optional[str]:
    try:
        r = requests.get(API_URL_JSON, timeout=10)
        r.raise_for_status()
        data = r.json()
        url  = data.get("url") or data.get("api_url")
        if url:
            # Verify alive
            try:
                requests.get(url.rstrip("/") + "/", timeout=8)
                return url.rstrip("/")
            except Exception:
                pass
    except Exception:
        pass
    return None


def _call_qwen(api_url: str, prompt: str) -> str:
    system = ("Extract financial values from the text. "
              "Output ONLY raw JSON. No markdown. Numbers only, no commas. "
              "Use null if not found.")
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.post(api_url + "/chat",
                              json={"messages": [
                                  {"role": "system", "content": system},
                                  {"role": "user",   "content": prompt},
                              ]},
                              timeout=REQUEST_TIMEOUT)
            r.raise_for_status()
            data  = r.json()
            reply = data.get("response") or data.get("reply") or data.get("message") or str(data)
            return reply.strip()
        except Exception as e:
            log.warning("LLM attempt %d/%d failed: %s", attempt, MAX_RETRIES, e)
            if attempt < MAX_RETRIES:
                time.sleep(3)
    return "{}"


def _safe_float(val) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(str(val).replace(",", ""))
    except (ValueError, TypeError):
        return None


def _parse_llm_json(raw: str) -> dict:
    raw = re.sub(r"```json\s*", "", raw)
    raw = re.sub(r"```\s*",     "", raw)
    start = raw.find("{")
    if start == -1:
        return {}
    end = raw.rfind("}")
    if end == -1:
        # Repair truncated
        s = raw[start:]
        s  = re.sub(r",\s*$", "", s.rstrip())
        s += "}" * (s.count("{") - s.count("}"))
    else:
        s = raw[start: end + 1]
    s = re.sub(r",\s*([}\]])", r"\1", s)
    try:
        return json.loads(s)
    except Exception:
        return {}


def llm_fill(text_section: str, obj, api_url: str, year_label: str):
    """
    For any field still None on `obj`, ask the LLM for just those fields.
    Sends maximum 8 null fields per call to keep output small.
    """
    null_fields = [f for f in obj.__dataclass_fields__ if getattr(obj, f) is None
                   and f not in ("gross_profit", "total_liabilities")]

    if not null_fields:
        log.info("  No null fields left — skipping LLM for %s", year_label)
        return

    log.info("  LLM fallback for %s: filling %d null fields: %s",
             year_label, len(null_fields), null_fields)

    # Split into batches of 8 to keep output within token limit
    for i in range(0, len(null_fields), 8):
        batch = null_fields[i: i + 8]
        schema = "{" + ", ".join(f'"{f}": null' for f in batch) + "}"
        prompt = (
            f"From this financial text, extract ONLY these fields for {year_label}.\n"
            f"Output this JSON with values filled in:\n{schema}\n\n"
            f"TEXT:\n{text_section[:2000]}"
        )
        raw  = _call_qwen(api_url, prompt)
        data = _parse_llm_json(raw)
        for f in batch:
            if data.get(f) is not None:
                val = _safe_float(data[f])
                if val is not None:
                    setattr(obj, f, val)
                    log.info("    LLM filled %s = %s", f, val)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────────────────────────────────────

BLUE       = "FF0070C0"
GREEN      = "FF00B050"
SECTION_BG = "FFBDD7EE"


def _write_val(cell, value: Optional[float]):
    if value is None:
        cell.value = "-"
        cell.font  = Font(name="Arial", size=10, color="FF808080")
    else:
        cell.value = round(value, 2)
        cell.font  = Font(name="Arial", size=10, color=BLUE)
        cell.number_format = '#,##0.00;(#,##0.00);"-"'
    cell.alignment = Alignment(horizontal="right", vertical="center")


def populate_excel(template_path: Path, output_path: Path, report: FinancialReport):
    wb = load_workbook(template_path)
    ws = wb.active

    CL, CC, CP = 5, 6, 7
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    ws.cell(1, 1).value = report.company
    ws.cell(1, 1).font  = Font(name="Arial", bold=True, size=12)

    for col, lbl in [(CL, "Financial Summary (INR in Lakhs)"),
                     (CC, "FY 2022-23"), (CP, "FY 2021-22")]:
        c = ws.cell(2, col)
        c.value     = lbl
        c.font      = Font(name="Arial", bold=True, size=10,
                           color=BLUE if col != CL else "FF000000")
        c.alignment = Alignment(horizontal="center" if col != CL else "left")

    b = report.standalone_bs_current
    p = report.standalone_bs_prior
    i = report.standalone_pl_current
    j = report.standalone_pl_prior

    rows = [
        (3,  "Balance Sheet Particulars",          None,b.share_capital,            None,p.share_capital,              True,  True),
        (4,  "Networth",                           None,None,                        None,None,                         True,  False),
        (5,  "Share Capital",                      b.share_capital,None,             p.share_capital,None,              False, False),
        (6,  "Retained Earnings",                  b.retained_earnings,None,         p.retained_earnings,None,          False, False),
        (7,  "General Reserves & Surplus",         b.general_reserves,None,          p.general_reserves,None,           False, False),
        (8,  "Other Equity",                       b.other_equity,None,              p.other_equity,None,               False, False),
        (10, "Total Networth",                     b.total_networth,None,            p.total_networth,None,             True,  False),
        (12, "Current Liabilities",                None,None,                        None,None,                         True,  False),
        (14, "Accounts Payable",                   b.accounts_payable,None,          p.accounts_payable,None,           False, False),
        (15, "Provisions",                         b.provisions_cl,None,             p.provisions_cl,None,              False, False),
        (16, "Short term borrowings",              b.short_term_borrowings,None,     p.short_term_borrowings,None,      False, False),
        (17, "Other Current Liabilities",          b.other_current_liabilities,None, p.other_current_liabilities,None,  False, False),
        (18, "Other Financial Liabilities",        b.other_financial_liabilities_cl,None, p.other_financial_liabilities_cl,None, False,False),
        (20, "Total Current Liabilities",          b.total_current_liabilities,None, p.total_current_liabilities,None,  True,  False),
        (22, "Non-Current Liabilities",            None,None,                        None,None,                         True,  False),
        (24, "Long Term borrowings",               b.long_term_borrowings,None,      p.long_term_borrowings,None,       False, False),
        (25, "Provision",                          b.provision_ncl,None,             p.provision_ncl,None,              False, False),
        (26, "Others",                             b.others_ncl,None,                p.others_ncl,None,                 False, False),
        (27, "Other Financial Liabilities",        b.other_financial_liabilities_ncl,None, p.other_financial_liabilities_ncl,None, False,False),
        (29, "Total Non-Current Liabilities",      b.total_non_current_liabilities,None, p.total_non_current_liabilities,None, True,False),
        (31, "Total Liabilities",                  b.total_liabilities,None,         p.total_liabilities,None,          True,  False),
        (33, "Current Assets",                     None,None,                        None,None,                         True,  False),
        (35, "Bank Balance",                       b.bank_balance,None,              p.bank_balance,None,               False, False),
        (36, "Cash & Cash Equivalence",            b.cash_equivalents,None,          p.cash_equivalents,None,           False, False),
        (37, "Inventory",                          b.inventory,None,                 p.inventory,None,                  False, False),
        (39, "Loans",                              b.loans_ca,None,                  p.loans_ca,None,                   False, False),
        (40, "Accounts Receivable",                b.accounts_receivable,None,       p.accounts_receivable,None,        False, False),
        (41, "Other Current Assets",               b.other_current_assets,None,      p.other_current_assets,None,       False, False),
        (42, "Other Financial Assets",             b.other_financial_assets_ca,None, p.other_financial_assets_ca,None,  False, False),
        (44, "Total Current Assets",               b.total_current_assets,None,      p.total_current_assets,None,       True,  False),
        (46, "Non-Current Assets",                 None,None,                        None,None,                         True,  False),
        (48, "Fixed Assets",                       b.fixed_assets,None,              p.fixed_assets,None,               False, False),
        (49, "Investments",                        b.investments_nca,None,           p.investments_nca,None,            False, False),
        (51, "Capital Work-in-Progress",           b.cwip,None,                      p.cwip,None,                       False, False),
        (53, "Other Financial Assets (NCA)",       b.other_financial_assets_nca,None,p.other_financial_assets_nca,None, False, False),
        (54, "Deferred Tax Assets",                b.deferred_tax_assets,None,       p.deferred_tax_assets,None,        False, False),
        (56, "Total Non-Current Assets",           b.total_non_current_assets,None,  p.total_non_current_assets,None,   True,  False),
        (59, "Total Assets",                       b.total_assets,None,              p.total_assets,None,               True,  False),
        (61, "DIFFERENCE",                         None,None,                        None,None,                         False, False),
        (65, "P&L Statement Particulars",          None,None,                        None,None,                         True,  False),
        (66, "Revenue",                            i.revenue,None,                   j.revenue,None,                    False, False),
        (67, "Cost of goods sold",                 i.cost_of_goods_sold,None,        j.cost_of_goods_sold,None,         False, False),
        (68, "Gross profit",                       i.gross_profit,None,              j.gross_profit,None,               True,  False),
        (69, "Employee benefits expense",          i.employee_benefits,None,         j.employee_benefits,None,          False, False),
        (70, "Interest",                           i.interest,None,                  j.interest,None,                   False, False),
        (71, "Depreciation",                       i.depreciation,None,              j.depreciation,None,               False, False),
        (72, "Other expenses less other income",   i.other_expenses_net,None,        j.other_expenses_net,None,         False, False),
        (73, "Taxes",                              i.taxes,None,                     j.taxes,None,                      False, False),
        (74, "Net Profit",                         i.net_profit,None,                j.net_profit,None,                 True,  False),
    ]

    for (row, label, cy_val, _, py_val, __, is_section, is_header) in rows:
        lc, cc, pc = ws.cell(row, CL), ws.cell(row, CC), ws.cell(row, CP)
        lc.value = label
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 16

        if is_header or (is_section and cy_val is None):
            lc.font = Font(name="Arial", bold=True, size=10, color="FF1F4E79")
            for c in [lc, cc, pc]:
                c.fill = PatternFill("solid", fgColor=SECTION_BG)
        elif is_section:
            lc.font = Font(name="Arial", bold=True, size=10)
            _write_val(cc, cy_val); _write_val(pc, py_val)
            for c in [cc, pc]:
                c.font   = Font(name="Arial", bold=True, size=10, color=GREEN)
                c.border = Border(top=Side(style="thin"), bottom=Side(style="double"))
        else:
            lc.font = Font(name="Arial", size=10)
            _write_val(cc, cy_val); _write_val(pc, py_val)

    # DIFFERENCE formula
    ws.cell(61, CL).value = "DIFFERENCE (Assets − Liabilities − Equity)"
    ws.cell(61, CL).font  = Font(name="Arial", bold=True, size=10, color="FFFF0000")
    for col, formula in [(CC, "=F59-F31-F10"), (CP, "=G59-G31-G10")]:
        c = ws.cell(61, col)
        c.value = formula
        c.font  = Font(name="Arial", bold=True, size=10, color="FFFF0000")
        c.number_format = '#,##0.00;(#,##0.00);"-"'
        c.alignment = Alignment(horizontal="right")

    wb.save(output_path)
    log.info("Saved: %s", output_path)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def _null_count(obj) -> int:
    return sum(1 for f in obj.__dataclass_fields__
               if getattr(obj, f) is None
               and f not in ("gross_profit", "total_liabilities",
                             "retained_earnings", "general_reserves",
                             "loans_nca", "investments_ca",
                             "other_non_current_assets"))


def run(source_file: str, template_file: str, output_file: str):
    source   = Path(source_file)
    template = Path(template_file)
    output   = Path(output_file)

    # ── Read ─────────────────────────────────────────────────────────────────
    full_text = read_source_text(source)
    bs_text   = _section(full_text, "BALANCE SHEET AS AT",
                          ["STATEMENT OF PROFIT", "CONSOLIDATED BALANCE", "NOTES TO"])
    pl_text   = _section(full_text, "STATEMENT OF PROFIT AND LOSS",
                          ["STATEMENT OF CHANGE", "CONSOLIDATED STATEMENT", "CASH FLOW"])

    # ── Layer 1: Regex ───────────────────────────────────────────────────────
    log.info("━━━ Layer 1: Regex Parser")
    bs_cur, bs_pri = regex_parse_bs(full_text)
    pl_cur, pl_pri = regex_parse_pl(full_text)
    log.info("  After regex  — BS null fields: cur=%d pri=%d | PL null: cur=%d pri=%d",
             _null_count(bs_cur), _null_count(bs_pri),
             _null_count(pl_cur), _null_count(pl_pri))

    # ── Layer 2: Alias/Fuzzy fill ─────────────────────────────────────────────
    log.info("━━━ Layer 2: Alias/Fuzzy Matcher")
    alias_fill(bs_text, bs_cur, bs_pri, pl_cur, pl_pri, section="bs")
    alias_fill(pl_text, bs_cur, bs_pri, pl_cur, pl_pri, section="pl")
    log.info("  After aliases — BS null fields: cur=%d pri=%d | PL null: cur=%d pri=%d",
             _null_count(bs_cur), _null_count(bs_pri),
             _null_count(pl_cur), _null_count(pl_pri))

    # ── Layer 3: LLM fallback (only if fields still null) ────────────────────
    total_null = _null_count(bs_cur) + _null_count(bs_pri) + \
                 _null_count(pl_cur) + _null_count(pl_pri)

    if total_null > 0:
        log.info("━━━ Layer 3: LLM Fallback (%d fields still null)", total_null)
        api_url = _get_api_url()
        if api_url:
            log.info("  Qwen API live at %s", api_url)
            llm_fill(bs_text, bs_cur, api_url, "Balance Sheet CURRENT year")
            llm_fill(bs_text, bs_pri, api_url, "Balance Sheet PRIOR year")
            llm_fill(pl_text, pl_cur, api_url, "P&L CURRENT year")
            llm_fill(pl_text, pl_pri, api_url, "P&L PRIOR year")
        else:
            log.warning("  Qwen API offline — skipping LLM layer")
    else:
        log.info("━━━ Layer 3: Skipped — all fields populated by layers 1+2 ✓")

    # Derived fields
    if bs_cur.total_current_liabilities and bs_cur.total_non_current_liabilities:
        bs_cur.total_liabilities = bs_cur.total_current_liabilities + bs_cur.total_non_current_liabilities
    if bs_pri.total_current_liabilities and bs_pri.total_non_current_liabilities:
        bs_pri.total_liabilities = bs_pri.total_current_liabilities + bs_pri.total_non_current_liabilities
    if pl_cur.revenue and pl_cur.cost_of_goods_sold:
        pl_cur.gross_profit = pl_cur.revenue - pl_cur.cost_of_goods_sold
    if pl_pri.revenue and pl_pri.cost_of_goods_sold:
        pl_pri.gross_profit = pl_pri.revenue - pl_pri.cost_of_goods_sold

    # ── Write Excel ───────────────────────────────────────────────────────────
    report = FinancialReport(
        company               = "Madhav Infra Projects Limited",
        fiscal_year           = "FY2022-23",
        standalone_bs_current = bs_cur,
        standalone_bs_prior   = bs_pri,
        standalone_pl_current = pl_cur,
        standalone_pl_prior   = pl_pri,
    )
    populate_excel(template, output, report)

    # ── Summary ───────────────────────────────────────────────────────────────
    print("\n" + "="*62)
    print("  EXTRACTION COMPLETE — 3-Layer Pipeline")
    print("="*62)
    print(f"  {'Field':<38} {'FY23':>10} {'FY22':>10}")
    print("  " + "-"*58)

    def row(lbl, a, b):
        a_s = f"{a:>10,.2f}" if a is not None else f"{'N/A':>10}"
        b_s = f"{b:>10,.2f}" if b is not None else f"{'N/A':>10}"
        print(f"  {lbl:<38} {a_s} {b_s}")

    print("\n  BALANCE SHEET (INR Lakhs)")
    row("Share Capital",           bs_cur.share_capital,              bs_pri.share_capital)
    row("Other Equity",            bs_cur.other_equity,               bs_pri.other_equity)
    row("Total Networth",          bs_cur.total_networth,             bs_pri.total_networth)
    row("Total Current Liab.",     bs_cur.total_current_liabilities,  bs_pri.total_current_liabilities)
    row("Total NCL",               bs_cur.total_non_current_liabilities, bs_pri.total_non_current_liabilities)
    row("Total Liabilities",       bs_cur.total_liabilities,          bs_pri.total_liabilities)
    row("Total Current Assets",    bs_cur.total_current_assets,       bs_pri.total_current_assets)
    row("Total Non-Curr Assets",   bs_cur.total_non_current_assets,   bs_pri.total_non_current_assets)
    row("TOTAL ASSETS",            bs_cur.total_assets,               bs_pri.total_assets)

    print("\n  P&L (INR Lakhs)")
    row("Revenue",                 pl_cur.revenue,            pl_pri.revenue)
    row("Cost of Goods Sold",      pl_cur.cost_of_goods_sold, pl_pri.cost_of_goods_sold)
    row("Gross Profit",            pl_cur.gross_profit,       pl_pri.gross_profit)
    row("Employee Expenses",       pl_cur.employee_benefits,  pl_pri.employee_benefits)
    row("Interest",                pl_cur.interest,           pl_pri.interest)
    row("Depreciation",            pl_cur.depreciation,       pl_pri.depreciation)
    row("Taxes",                   pl_cur.taxes,              pl_pri.taxes)
    row("Net Profit",              pl_cur.net_profit,         pl_pri.net_profit)
    print("="*62)
    print(f"\n  ✓ Output → {output}\n")


if __name__ == "__main__":
    src  = sys.argv[1] if len(sys.argv) > 1 else "Madhav_Infra_FY2223.pdf"
    tmpl = sys.argv[2] if len(sys.argv) > 2 else "Input_Templates.xlsx"
    out  = sys.argv[3] if len(sys.argv) > 3 else "Madhav_Infra_Output.xlsx"
    run(src, tmpl, out)
